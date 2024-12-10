from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
from selenium.webdriver.remote.webelement import WebElement
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from typing import List, Dict, Optional, Union, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import logging
import time
import os

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('form_automation.log'),
        logging.StreamHandler()
    ]
)

class FormField:
    """Class to represent a form field"""
    def __init__(self, name: str, field_type: str, label: str, required: bool = False):
        self.name = name
        self.field_type = field_type
        self.label = label
        self.required = required

    def to_dict(self) -> Dict[str, Union[str, bool]]:
        return {
            'name': self.name,
            'type': self.field_type,
            'label': self.label,
            'required': self.required
        }

class FormAnalyzer:
    def __init__(self, url: str):
        """Initialize the form analyzer"""
        self.url = url
        self.driver: Optional[webdriver.Edge] = None
        self.wait: Optional[WebDriverWait] = None
        self.form_fields: List[FormField] = []
        
    def setup_driver(self) -> None:
        """Initialize Edge driver with custom options"""
        try:
            edge_options = Options()
            edge_options.add_argument('--start-maximized')
            edge_options.add_argument('--disable-blink-features=AutomationControlled')
            edge_options.add_argument('--disable-notifications')
            edge_options.add_argument('--inprivate')
            edge_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            edge_options.add_experimental_option('useAutomationExtension', False)
            
            service = Service(EdgeChromiumDriverManager().install())
            self.driver = webdriver.Edge(service=service, options=edge_options)
            self.wait = WebDriverWait(self.driver, 10)
            logging.info("Browser setup completed successfully")
            
        except Exception as e:
            logging.error(f"Error setting up browser: {e}")
            if self.driver:
                self.driver.quit()
            self.driver = None
            raise

    def find_label_text(self, field: WebElement) -> str:
        """Find label text for a form field using multiple strategies"""
        if not self.driver:
            return ""
            
        field_id = field.get_attribute('id')
        if field_id:
            try:
                label_elem = self.driver.find_element(By.CSS_SELECTOR, f"label[for='{field_id}']")
                if label_elem:
                    return label_elem.text.strip()
            except NoSuchElementException:
                pass

        try:
            # Look for label as ancestor
            label_elem = field.find_element(By.XPATH, "./ancestor::label")
            if label_elem:
                return label_elem.text.strip()
        except NoSuchElementException:
            pass

        try:
            # Look for label as sibling
            label_elem = field.find_element(By.XPATH, "../label")
            if label_elem:
                return label_elem.text.strip()
        except NoSuchElementException:
            pass

        return ""

    def get_field_info(self, field: WebElement) -> Optional[FormField]:
        """Extract information about a form field"""
        try:
            field_type = field.get_attribute('type') or 'text'
            field_id = field.get_attribute('id') or ''
            field_name = field.get_attribute('name') or field_id
            field_placeholder = field.get_attribute('placeholder') or ''
            
            # Get label text
            label = self.find_label_text(field)
            
            # Use the best available name for the field
            field_label = label or field_placeholder or field_name or field_id
            if not field_label:
                return None
                
            required = field.get_attribute('required') is not None or field.get_attribute('aria-required') == 'true'
            
            return FormField(
                name=field_name,
                field_type=field_type,
                label=field_label,
                required=required
            )
            
        except Exception as e:
            logging.error(f"Error getting field info: {e}")
            return None

    def analyze_form(self) -> None:
        """Analyze form fields"""
        try:
            self.setup_driver()
            if not self.driver or not self.wait:
                return
            
            # Navigate to the form URL
            self.driver.get(self.url)
            time.sleep(2)  # Wait for page load
            
            # Find all input fields
            input_elements = self.wait.until(EC.presence_of_all_elements_located(
                (By.XPATH, "//input[not(@type='hidden')] | //select | //textarea")
            ))
            
            for field in input_elements:
                field_info = self.get_field_info(field)
                if field_info:
                    self.form_fields.append(field_info)
                    logging.info(f"Found field: {field_info.label} ({field_info.field_type})")
            
            logging.info(f"Found {len(self.form_fields)} form fields")
            
        except Exception as e:
            logging.error(f"Error analyzing form: {e}")
        finally:
            if self.driver:
                self.driver.quit()

    def create_excel_template(self) -> None:
        """Create Excel template with form fields"""
        if not self.form_fields:
            logging.warning("No form fields found")
            return
            
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            if not ws:
                raise ValueError("Could not get active worksheet")

            # Add headers
            for idx, field in enumerate(self.form_fields, start=1):
                header = field.label
                if field.required:
                    header += " *"
                ws.cell(row=1, column=idx, value=header)

            # Add data validation for email fields
            for idx, field in enumerate(self.form_fields, start=1):
                if field.field_type == 'email':
                    # Add email validation for all rows in this column
                    col_letter = get_column_letter(idx)
                    dv = DataValidation(
                        type="custom",
                        formula1='=ISNUMBER(FIND("@",INDIRECT(ADDRESS(ROW(),COLUMN()))))',
                        allow_blank=True
                    )
                    dv.error = 'Please enter a valid email address'
                    dv.errorTitle = 'Invalid Email'
                    ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}1000")

            # Save the workbook
            output_file = 'form_template.xlsx'
            wb.save(output_file)
            
            logging.info(f"Created Excel template: {output_file}")
            print(f"\nCreated Excel template: {output_file}")
            print("* indicates required fields")
            
        except Exception as e:
            logging.error(f"Error creating Excel template: {e}")
            raise

def main() -> None:
    """Main function to run the form analyzer"""
    try:
        url = input("Enter the form URL: ")
        print("\nAnalyzing form fields...")
        
        analyzer = FormAnalyzer(url)
        analyzer.analyze_form()
        
        if analyzer.form_fields:
            print("\nFound the following fields:")
            for field in analyzer.form_fields:
                required = "*" if field.required else " "
                print(f"{required} {field.label} ({field.field_type})")
            
            analyzer.create_excel_template()
            print("\nTemplate created successfully! You can now:")
            print("1. Open form_template.xlsx")
            print("2. Fill in the required information")
            print("3. Run form_filler.py to submit the form")
        else:
            print("\nNo form fields found. Please check the URL and try again.")
        
    except Exception as e:
        logging.error(f"Error in main: {e}")
        print("\nAn error occurred. Please check form_automation.log for details.")

if __name__ == "__main__":
    main()
